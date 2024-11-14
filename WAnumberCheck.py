from openpyxl import load_workbook

def get_mobile_numbers(file_path='testWA.xlsx'):
    # Load the workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Read mobile numbers from the first column and return as a list
    mobile_numbers = []
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        mobile_number = row[0]
        if mobile_number:
            mobile_numbers.append(str(mobile_number))
    
    return mobile_numbers

def log_no_whatsapp(mobile_number, file_path='mobile_numbers.xlsx'):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    row = sheet.max_row + 1
    sheet[f"A{row}"] = mobile_number
    sheet[f"B{row}"] = "No WhatsApp"

    workbook.save(file_path)
